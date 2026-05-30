#!/usr/bin/env python3
"""Extract Colab links from syllabus xlsx and audit HTTP accessibility."""
from __future__ import annotations

import argparse
import re
import sys
import urllib.request
from pathlib import Path

try:
    import openpyxl
    import yaml
except ImportError:
    sys.exit("Requires: pip install openpyxl pyyaml")

COLAB_RE = re.compile(r"https://colab\.research\.google\.com/[^\s\"']+")


def extract_from_syllabus(path: Path) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    sessions = []
    for r in range(2, ws.max_row + 1):
        num = ws.cell(r, 1).value
        title = ws.cell(r, 2).value
        material = ws.cell(r, 4).value or ""
        if num is None or not title:
            continue
        links = list(dict.fromkeys(COLAB_RE.findall(str(material))))
        sessions.append({
            "session": int(num),
            "title": str(title).strip(),
            "colab_links": links,
            "audit_status": "pending",
            "quiz_notes": "",
            "feedback_signals": [],
        })
    return sessions


def check_url(url: str, timeout: int = 10) -> str:
    try:
        req = urllib.request.Request(url, method="HEAD")
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return "ok" if resp.status < 400 else f"http_{resp.status}"
    except Exception as e:
        return f"error:{type(e).__name__}"


def audit_sessions(sessions: list[dict]) -> None:
    for s in sessions:
        if not s["colab_links"]:
            s["audit_status"] = "no_colab"
            continue
        results = [check_url(u) for u in s["colab_links"]]
        if all(r == "ok" for r in results):
            s["audit_status"] = "ok"
        else:
            s["audit_status"] = "needs_fix"
            s["link_checks"] = dict(zip(s["colab_links"], results))


def main() -> None:
    p = argparse.ArgumentParser()
    p.add_argument("--syllabus", type=Path, required=True)
    p.add_argument("--output", type=Path, required=True)
    p.add_argument("--skip-http-check", action="store_true")
    args = p.parse_args()

    sessions = extract_from_syllabus(args.syllabus)
    if not args.skip_http_check:
        audit_sessions(sessions)

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(yaml.dump({"sessions": sessions}, sort_keys=False, allow_unicode=True))

    print(f"Wrote {len(sessions)} sessions to {args.output}")
    for s in sessions:
        n = len(s["colab_links"])
        print(f"  S{s['session']}: {n} colab(s) — {s['audit_status']}")


if __name__ == "__main__":
    main()
