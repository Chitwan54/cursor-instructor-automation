#!/usr/bin/env python3
"""
Extract Scaler problem/test IDs from assignment tracker files.

Supports:
  - Google Sheets exported as .xlsx (hyperlinks preserved)
  - Google Docs exported as .docx (hyperlinks in document.xml.rels)

Usage:
  python extract_ids.py --input tracker.xlsx [--output extracted.json]
  python extract_ids.py --docx assessments.docx [--output extracted.json]
  python extract_ids.py --input tracker.xlsx --tab "NumPy assessments"
"""
from __future__ import annotations

import argparse
import json
import re
import sys
import zipfile
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required: pip install openpyxl")

PROBLEM_RE = re.compile(r"/hire/test/problem/(\d+)")
TEST_RE = re.compile(r"/hire/test/(\d+)/?")


def ids_from_url(url: str | None) -> list[tuple[str, str]]:
    """Return [(id, type), ...] from a hyperlink target."""
    if not url:
        return []
    out: list[tuple[str, str]] = []
    for m in PROBLEM_RE.finditer(url):
        out.append((m.group(1), "problem"))
    if not out:
        for m in TEST_RE.finditer(url):
            out.append((m.group(1), "test"))
    return out


def extract_xlsx(path: Path, tab: str | None = None) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    tabs = [tab] if tab else wb.sheetnames
    result: dict = {"source": str(path), "tabs": {}}

    for name in tabs:
        if name not in wb.sheetnames:
            print(f"Warning: tab '{name}' not found", file=sys.stderr)
            continue
        ws = wb[name]
        cur_section: str | None = None
        sections: dict[str, list[dict]] = {}

        for r in range(1, ws.max_row + 1):
            # Section header: lecture number or topic in cols A/B
            a, b = ws.cell(r, 1).value, ws.cell(r, 2).value
            if b and str(b).strip():
                label = f"{a} - {b}".strip() if a else str(b).strip()
                cur_section = label
            elif a and isinstance(a, (int, float)) and 1 <= a <= 20:
                cur_section = f"Lecture {int(a)}"

            if not cur_section:
                continue

            for c in range(1, min(ws.max_column + 1, 30)):
                cell = ws.cell(r, c)
                targets = []
                if cell.hyperlink and cell.hyperlink.target:
                    targets.append(cell.hyperlink.target)
                if cell.value and isinstance(cell.value, str) and "scaler.com" in cell.value:
                    targets.append(cell.value)

                for target in targets:
                    for pid, id_type in ids_from_url(target):
                        name_text = str(cell.value or "").strip()
                        if _should_skip(name_text):
                            continue
                        sections.setdefault(cur_section, []).append({
                            "id": pid,
                            "type": id_type,
                            "name": name_text,
                            "row": r,
                            "col": c,
                        })

        # Deduplicate IDs per section (preserve order)
        for sec in sections:
            seen: set[str] = set()
            deduped = []
            for item in sections[sec]:
                if item["id"] not in seen:
                    seen.add(item["id"])
                    deduped.append(item)
            sections[sec] = deduped

        result["tabs"][name] = {
            "sections": sections,
            "total_ids": sum(len(v) for v in sections.values()),
        }

    return result


def extract_docx(path: Path) -> dict:
    with zipfile.ZipFile(path) as z:
        rels = z.read("word/_rels/document.xml.rels").decode("utf-8", errors="ignore")
        doc = z.read("word/document.xml").decode("utf-8", errors="ignore")

    rel_map = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels))
    current: str | None = None
    sections: dict[str, list[dict]] = {}

    for para in re.split(r"</w:p>", doc):
        text = re.sub(r"<[^>]+>", "", para)
        text = re.sub(r"\s+", " ", text).strip()
        if not text:
            continue
        # Treat lines with " - N" or topic headers as section boundaries
        if re.search(r"\b-\s*\d+\b", text) or len(text) < 80 and text[0].isupper():
            current = text

        for m in re.finditer(r'<w:hyperlink[^>]*r:id="(rId\d+)"', para):
            target = rel_map.get(m.group(1), "")
            for pid, id_type in ids_from_url(target):
                sec = current or "default"
                sections.setdefault(sec, []).append({
                    "id": pid,
                    "type": id_type,
                    "name": text[:120],
                })

    for sec in sections:
        seen: set[str] = set()
        deduped = [x for x in sections[sec] if x["id"] not in seen and not seen.add(x["id"])]

    return {
        "source": str(path),
        "tabs": {"docx": {"sections": sections, "total_ids": sum(len(v) for v in sections.values())}},
    }


def _should_skip(name: str) -> bool:
    lower = name.lower()
    skip_phrases = ("not in use", "not covered", "deprecated", "do not use")
    return any(p in lower for p in skip_phrases)


def print_summary(data: dict) -> None:
    for tab, info in data.get("tabs", {}).items():
        print(f"\n=== {tab} ({info['total_ids']} IDs) ===")
        for sec, items in info["sections"].items():
            ids = [x["id"] for x in items]
            print(f"  {sec}: {len(ids)} IDs — {', '.join(ids[:8])}{'...' if len(ids) > 8 else ''}")


def main() -> None:
    p = argparse.ArgumentParser(description="Extract Scaler problem IDs from assignment trackers")
    p.add_argument("--input", type=Path, help="Input .xlsx file")
    p.add_argument("--docx", type=Path, help="Input .docx file")
    p.add_argument("--tab", help="Specific xlsx tab name")
    p.add_argument("--output", "-o", type=Path, help="Write JSON output")
    p.add_argument("--summary", action="store_true", default=True, help="Print summary (default)")
    args = p.parse_args()

    if args.docx:
        data = extract_docx(args.docx)
    elif args.input:
        data = extract_xlsx(args.input, args.tab)
    else:
        p.error("Provide --input or --docx")

    if args.output:
        args.output.write_text(json.dumps(data, indent=2))
        print(f"Wrote {args.output}")

    if args.summary:
        print_summary(data)


if __name__ == "__main__":
    main()
